import json
import os
import openpyxl
import re

workspace_dir = "/Users/diegozolhos/Projects/AuxiliarBNCC"
raw_dir = os.path.join(workspace_dir, "data", "raw")
out_dir = os.path.join(workspace_dir, "public")

file_completa = os.path.join(raw_dir, "BNCC_Completa.json")
file_computacao = os.path.join(raw_dir, "BNCC_Computacao_Rica.json")
file_xlsx = os.path.join(raw_dir, "BNCC_COMPUTAÇÃO_BY_ZOLHOS.xlsx")

out_completa = os.path.join(out_dir, "BNCC_Completa_Otimizada.json")
out_computacao = os.path.join(out_dir, "BNCC_Computacao_Otimizada.json")

print("Loading original datasets...")
with open(file_completa, 'r', encoding='utf-8') as f:
    data_completa = json.load(f)

with open(file_computacao, 'r', encoding='utf-8') as f:
    data_computacao = json.load(f)

# Helper function to title case axes
def format_eixo(eixo_str):
    if not eixo_str:
        return None
    eixo_clean = eixo_str.strip().upper()
    if eixo_clean == "PENSAMENTO COMPUTACIONAL":
        return "Pensamento Computacional"
    elif eixo_clean == "MUNDO DIGITAL":
        return "Mundo Digital"
    elif eixo_clean == "CULTURA DIGITAL":
        return "Cultura Digital"
    return eixo_str.strip()

# Helper function to normalize ano_faixa
def normalize_ano_faixa(anos, etapa):
    if not anos:
        return []
    res = []
    for a in anos:
        a_clean = a.strip().upper()
        if a_clean == "INFANTIL":
            res.append("Educação Infantil")
        elif a_clean == "POR_ETAPA_1_AO_5_º":
            res.extend(["1º", "2º", "3º", "4º", "5º"])
        elif a_clean == "POR_ETAPA_6_AO_9_º":
            res.extend(["6º", "7º", "8º", "9º"])
        elif a_clean == "ENSINO_MÉDIO" or a_clean == "ENSINO_MEDIO":
            res.extend(["1º EM", "2º EM", "3º EM"])
        else:
            res.append(a.strip())
    return list(sorted(set(res), key=lambda x: (
        0 if "Infantil" in x else (
            1 if "EM" in x else (
                2 if x.endswith("º") else 3
            )
        ), x
    )))

# Parse competencies from Excel
print("Parsing competencies from Excel...")
wb = openpyxl.load_workbook(file_xlsx, read_only=True)

def parse_items(text):
    if not text:
        return []
    matches = re.findall(r'(?:^|\n)\s*(\d+[\.\)]\s+.*?)(?=\n\s*\d+[\.\)]|\Z)', text, flags=re.DOTALL)
    cleaned = []
    for m in matches:
        c = re.sub(r'\s+', ' ', m.strip())
        cleaned.append(c)
    return cleaned

premissas_infantil = []
competencias_fundamental = []
competencias_medio = []

if "INFANTIL_COMPETÊNCIAS" in wb.sheetnames:
    text = wb["INFANTIL_COMPETÊNCIAS"].cell(row=1, column=1).value
    premissas_infantil = parse_items(text)
    print(f"  Extracted {len(premissas_infantil)} premissas for Ed. Infantil.")

if "FUNDAMENTAL_FUNDAMENTAL_COMPETÊ" in wb.sheetnames:
    text = wb["FUNDAMENTAL_FUNDAMENTAL_COMPETÊ"].cell(row=1, column=1).value
    competencias_fundamental = parse_items(text)
    print(f"  Extracted {len(competencias_fundamental)} competencies for Ensino Fundamental.")

if "ENSINO_MÉDIO_COMPETÊNCIAS" in wb.sheetnames:
    text = wb["ENSINO_MÉDIO_COMPETÊNCIAS"].cell(row=1, column=1).value
    competencias_medio = parse_items(text)
    print(f"  Extracted {len(competencias_medio)} competencies for Ensino Médio.")

wb.close()

# 1. OPTIMIZE BNCC COMPLETA
print("Optimizing BNCC Completa...")
opt_completa_habs = []
for h in data_completa.get('habilidades', []):
    obj_conh = h.get('objetos_conhecimento')
    if isinstance(obj_conh, list):
        objs_list = [x.strip() for x in obj_conh if isinstance(x, str)]
    elif isinstance(obj_conh, str):
        objs_list = [obj_conh.strip()]
    else:
        objs_list = []

    opt_h = {
        "codigo": h.get('codigo'),
        "id_habilidade": h.get('id_habilidade') or h.get('codigo'),
        "etapa": "Ensino Fundamental",
        "componente": h.get('componente'),
        "ano_faixa": h.get('ano_faixa', []),
        "eixo": format_eixo(h.get('eixo')),
        "unidade_tematica": h.get('unidade_tematica'),
        "campos_atuacao": h.get('campos_atuacao'),
        "praticas_linguagem": h.get('praticas_linguagem'),
        "objetos_conhecimento": objs_list,
        "descricao": h.get('descricao'),
        "competencia_especifica": None,
        "comentario": h.get('comentario'),
        "possibilidades_curriculo": h.get('possibilidades_curriculo'),
        "explicacao_pedagogica": None,
        "exemplos": None
    }
    opt_completa_habs.append(opt_h)

data_completa_opt = {
    "competencias_gerais": data_completa.get("competencias_gerais", []),
    "competencias_especificas": data_completa.get("competencias_especificas", []),
    "habilidades": opt_completa_habs
}

with open(out_completa, 'w', encoding='utf-8') as f:
    json.dump(data_completa_opt, f, ensure_ascii=False, indent=2)
print(f"Saved optimized BNCC Completa to {out_completa}")

# 2. OPTIMIZE BNCC COMPUTACAO
print("Optimizing BNCC Computacao...")
opt_computacao_habs = []
seen_codes = set()

em_eixo_mapping = {
    "EM13CO01": "Pensamento Computacional",
    "EM13CO02": "Pensamento Computacional",
    "EM13CO03": "Pensamento Computacional",
    "EM13CO04": "Pensamento Computacional",
    "EM13CO05": "Pensamento Computacional",
    "EM13CO06": "Pensamento Computacional",
    "EM13CO07": "Mundo Digital",
    "EM13CO08": "Mundo Digital",
    "EM13CO09": "Cultura Digital",
    "EM13CO10": "Pensamento Computacional",
    "EM13CO11": "Pensamento Computacional",
    "EM13CO12": "Mundo Digital",
    "EM13CO13": "Mundo Digital",
    "EM13CO14": "Cultura Digital",
    "EM13CO15": "Cultura Digital",
    "EM13CO16": "Mundo Digital",
    "EM13CO17": "Cultura Digital",
    "EM13CO18": "Cultura Digital",
    "EM13CO19": "Cultura Digital",
    "EM13CO20": "Cultura Digital",
    "EM13CO21": "Cultura Digital",
    "EM13CO22": "Cultura Digital",
    "EM13CO23": "Cultura Digital",
    "EM13CO24": "Cultura Digital",
    "EM13CO25": "Cultura Digital",
    "EM13CO26": "Cultura Digital"
}

for h in data_computacao.get('habilidades', []):
    code = h.get('codigo')
    if not code:
        continue
    
    if code in seen_codes:
        continue
    seen_codes.add(code)

    etapa_orig = h.get('etapa')
    if etapa_orig == 'Infantil':
        etapa = "Educação Infantil"
    elif etapa_orig == 'Fundamental':
        etapa = "Ensino Fundamental"
    elif etapa_orig == 'Médio' or etapa_orig == 'ENSINO_MÉDIO':
        etapa = "Ensino Médio"
    else:
        etapa = etapa_orig

    obj_conh_orig = h.get('objeto_conhecimento') or {}
    macro = obj_conh_orig.get('macro_categoria')
    sub = obj_conh_orig.get('sub_area')
    
    objs_list = []
    def fix_typos(s):
        if not s:
            return s
        s = s.strip()
        s = s.replace("responsabilidad e", "responsabilidade")
        s = s.replace("responsabilidade no uso da tecnologia computacional", "responsabilidade no uso de tecnologia computacional")
        if s == "Segurança e responsabilidade no uso da tecnologia":
            return "Segurança e responsabilidade no uso de tecnologia computacional"
        return s

    macro = fix_typos(macro)
    sub = fix_typos(sub)

    if macro:
        objs_list.append(macro)
    if sub and sub != macro:
        objs_list.append(sub)

    hab_orig = h.get('habilidade') or {}
    descricao = hab_orig.get('diretriz_especifica')
    macro_bloco = hab_orig.get('objetivo_macro_bloco')

    eixo_orig = h.get('eixo')
    competencia_especifica = None
    
    if code.startswith("EM") or etapa == "Ensino Médio":
        competencia_especifica = eixo_orig
        eixo = em_eixo_mapping.get(code, "Cultura Digital")
    else:
        eixo = format_eixo(eixo_orig)

    if macro_bloco:
        competencia_especifica = macro_bloco

    opt_h = {
        "codigo": code,
        "id_habilidade": code,
        "etapa": etapa,
        "componente": "Computação",
        "ano_faixa": normalize_ano_faixa(h.get('ano_faixa', []), etapa),
        "eixo": eixo,
        "unidade_tematica": None,
        "campos_atuacao": None,
        "praticas_linguagem": None,
        "objetos_conhecimento": objs_list,
        "descricao": descricao,
        "competencia_especifica": competencia_especifica,
        "comentario": None,
        "possibilidades_curriculo": None,
        "explicacao_pedagogica": h.get('explicacao_pedagogica'),
        "exemplos": h.get('exemplos')
    }
    opt_computacao_habs.append(opt_h)

data_computacao_opt = {
    "premissas_infantil": premissas_infantil,
    "competencias_fundamental": competencias_fundamental,
    "competencias_medio": competencias_medio,
    "habilidades": opt_computacao_habs
}

with open(out_computacao, 'w', encoding='utf-8') as f:
    json.dump(data_computacao_opt, f, ensure_ascii=False, indent=2)
print(f"Saved optimized BNCC Computacao to {out_computacao}")

print("Verification check:")
print(f"Total optimized BNCC Completa habilidades: {len(opt_completa_habs)}")
print(f"Total optimized BNCC Computacao habilidades: {len(opt_computacao_habs)}")
print(f"Integrated Infantil Premissas: {len(premissas_infantil)}")
print(f"Integrated Fundamental Competencies: {len(competencias_fundamental)}")
print(f"Integrated Medio Competencies: {len(competencias_medio)}")
